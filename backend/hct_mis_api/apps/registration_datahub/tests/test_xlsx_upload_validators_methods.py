import operator
from unittest import TestCase, mock

import openpyxl
from django.conf import settings
from django.core.management import call_command
from openpyxl_image_loader import SheetImageLoader


class TestXLSXValidatorsMethods(TestCase):
    FILES_DIR_PATH = f"{settings.PROJECT_ROOT}/apps/registration_datahub/tests/test_file"

    def setUp(self) -> None:
        from hct_mis_api.apps.registration_datahub.validators import UploadXLSXValidator

        self.UploadXLSXValidator = UploadXLSXValidator

        call_command("loadflexfieldsattributes")

    def test_geolocation_validator(self):
        # test correct values:
        correct_values = (
            "1.1, 1.1",
            "0.0, 0.0",
            "54.1234252, 67.535232",
        )
        for value in correct_values:
            self.assertTrue(self.UploadXLSXValidator.geolocation_validator(value, "hh_geopoint_h_c"))

        # test incorrect values:
        incorrect_values = (
            "1, 1, 1, 1",
            "0, 0",
            "52.124.124, 1241.242",
            "24.121a, bcd421.222",
        )

        for value in incorrect_values:
            self.assertFalse(self.UploadXLSXValidator.geolocation_validator(value, "hh_geopoint_h_c"))

    def test_date_validator(self):
        # test correct values:
        correct_values = (
            "01-03-1994",
            "1-3-1994",
            "27-12-2020",
            "27/12/2020",
            "27.12.2020",
            "27.12.2020",
        )
        for value in correct_values:
            self.assertTrue(self.UploadXLSXValidator.date_validator(value, "birth_date_i_c"))

        # test incorrect values:
        incorrect_values = (
            "13-13-1994",
            "213.22.2020",
            "qwerty",
            "24",
            "-24",
        )

        for value in incorrect_values:
            self.assertFalse(self.UploadXLSXValidator.date_validator(value, "birth_date_i_c"))

    def test_integer_validator(self):
        # test correct values:
        correct_values = (
            "12",
            "0",
            0,
            12,
            12345,
            -12,
        )
        for value in correct_values:
            self.assertTrue(self.UploadXLSXValidator.integer_validator(value, "size_h_c"))

        # test incorrect values:
        incorrect_values = (
            "13-13-1994",
            "213.22.2020",
            "qwerty",
            # 12.2345,
            "12,242",
        )

        for value in incorrect_values:
            self.assertFalse(self.UploadXLSXValidator.integer_validator(value, "size_h_c"))

    def test_phone_validator(self):
        # test correct values:
        correct_values = (
            "+1-202-555-0172",
            "+44 1632 960852",
            "+1-613-555-0182",
            "+61 1900 654 321",
            "+36 55 979 922",
            "+353 20 915 8245",
            "+48 69 563 7300",
        )
        for value in correct_values:
            self.assertTrue(self.UploadXLSXValidator.phone_validator(value, "phone_no_i_c"))

        # test incorrect values:
        incorrect_values = (
            "123 123 123",  # no region code
            "13-13-1994",
            "213.22.2020",
            "qwerty",
            12.2345,
            "12,242",
            "12",
            12,
        )

        for value in incorrect_values:
            self.assertFalse(self.UploadXLSXValidator.phone_validator(value, "phone_no_i_c"))

    def test_choice_validator(self):
        test_correct_values = (("REFUGEE", "residence_status_h_c"),)
        test_incorrect_values = (
            ("YES", "work_status"),
            ("OTHER", "work_status"),
            ("Hearing Problems", "disability"),
            ("Option 37", "assistance_type_h_f"),
        )
        for value, header in test_correct_values:
            self.assertTrue(self.UploadXLSXValidator.choice_validator(value, header))

        for value, header in test_incorrect_values:
            self.assertFalse(self.UploadXLSXValidator.choice_validator(value, header))

    def test_rows_validator_too_many_head_of_households(self):
        wb = openpyxl.load_workbook(f"{self.FILES_DIR_PATH}/error-xlsx.xlsx", data_only=True,)
        result = self.UploadXLSXValidator.rows_validator(wb["Individuals"])
        expected = [
            {
                "row_number": 0,
                "header": "relationship_i_c",
                "message": "Sheet: Individuals, There are multiple head of " "households for household with id: 3",
            }
        ]

        self.assertEqual(expected, result)

    def test_rows_validator(self):

        wb = openpyxl.load_workbook(f"{self.FILES_DIR_PATH}/invalid_rows.xlsx", data_only=True,)

        wb_valid = openpyxl.load_workbook(f"{self.FILES_DIR_PATH}/new_reg_data_import.xlsx", data_only=True,)
        self.UploadXLSXValidator.image_loader = SheetImageLoader(wb["Individuals"])

        sheets_and_expected_values = (
            (
                wb["Households"],
                [
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 1 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 3,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 1, Option 2, Option "
                        "3 for type select many of field assistance_type_h_f",
                        "row_number": 4,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 13 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 5,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 3 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 6,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 1, Option 3 for type "
                        "select many of field assistance_type_h_f",
                        "row_number": 7,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 2, Option 3 for type "
                        "select many of field assistance_type_h_f",
                        "row_number": 8,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 2 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 9,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 1, Option 2, Option "
                        "4 for type select many of field assistance_type_h_f",
                        "row_number": 10,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 4 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 11,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 5 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 12,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 1, Option 4 for type "
                        "select many of field assistance_type_h_f",
                        "row_number": 13,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 2, Option 4 for type "
                        "select many of field assistance_type_h_f",
                        "row_number": 14,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 3 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 15,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 1, Option 2, Option "
                        "5 for type select many of field assistance_type_h_f",
                        "row_number": 16,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 6 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 17,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 7 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 18,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 1, Option 5 for type "
                        "select many of field assistance_type_h_f",
                        "row_number": 19,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 2, Option 5 for type "
                        "select many of field assistance_type_h_f",
                        "row_number": 20,
                    },
                    {
                        "header": "assistance_type_h_f",
                        "message": "Sheet: Households, Unexpected value: Option 4 for type select "
                        "many of field assistance_type_h_f",
                        "row_number": 21,
                    },
                ],
            ),
            (wb["Individuals"], [],),
            (wb_valid["Households"], [],),
            (wb_valid["Individuals"], [],),
        )

        for sheet, expected_values in sheets_and_expected_values:
            validator_class = self.UploadXLSXValidator()
            validator_class.image_loader = SheetImageLoader(sheet)
            result = validator_class.rows_validator(sheet)
            self.assertEqual(result, expected_values)

    def test_validate_file_extension(self):
        files_to_test = (
            (
                f"{self.FILES_DIR_PATH}/" f"image.png",
                [{"row_number": 1, "message": "Only .xlsx files are accepted for import"}],
            ),
            (f"{self.FILES_DIR_PATH}/" f"not_excel_file.xlsx", [{"row_number": 1, "message": "Invalid .xlsx file"}],),
        )

        for file_path, expected_values in files_to_test:
            with open(file_path, "rb") as file:
                result = self.UploadXLSXValidator.validate_file_extension(file=file)
                self.assertEqual(result[0]["row_number"], expected_values[0]["row_number"])
                self.assertEqual(result[0]["message"], expected_values[0]["message"])

    def test_validate_file_with_template(self):
        invalid_cols_file_path = f"{self.FILES_DIR_PATH}/new_reg_data_import.xlsx"
        with open(invalid_cols_file_path, "rb") as file:
            errors = self.UploadXLSXValidator.validate_file_with_template(file=file)
            errors.sort(key=operator.itemgetter("row_number", "header"))
            self.assertEqual(errors, [])

    def test_required_validator(self):
        with mock.patch.dict(
            "registration_datahub.validators.UploadXLSXValidator.ALL_FIELDS", {"test": {"required": True}}, clear=True,
        ):
            result = self.UploadXLSXValidator.required_validator(value="tak", header="test")
            self.assertTrue(result)

        with mock.patch.dict(
            "registration_datahub.validators.UploadXLSXValidator.ALL_FIELDS", {"test": {"required": True}}, clear=True,
        ):
            result = self.UploadXLSXValidator.required_validator(value="", header="test")
            self.assertFalse(result)

        with mock.patch.dict(
            "registration_datahub.validators.UploadXLSXValidator.ALL_FIELDS", {"test": {"required": False}}, clear=True,
        ):
            result = self.UploadXLSXValidator.required_validator(value="", header="test")
            self.assertTrue(result)
