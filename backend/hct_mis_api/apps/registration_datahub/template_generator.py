from typing import Dict, List, Optional, Tuple

import openpyxl

from hct_mis_api.apps.core.field_attributes.core_fields_attributes import FieldFactory
from hct_mis_api.apps.core.field_attributes.fields_types import Scope
from hct_mis_api.apps.core.utils import serialize_flex_attributes
from hct_mis_api.apps.geo.models import Area


class TemplateFileGenerator:
    @classmethod
    def _create_workbook(cls) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws_households = wb.active
        ws_households.title = "Households"
        wb.create_sheet("Individuals")
        wb.create_sheet("Choices")

        return wb

    @classmethod
    def _handle_choices(cls, fields: Dict) -> List[List[str]]:
        rows: list[list[str]] = [["Field Name", "Label", "Value to be used in template"]]

        for field_name, field_value in fields.items():
            is_admin_level = field_name in ("admin1_h_c", "admin2_h_c")
            choices = field_value["choices"]
            if is_admin_level:
                choices = Area.get_admin_areas_as_choices(field_name[-5])
            if choices:
                for choice in field_value["choices"]:
                    row = [
                        field_name,
                        str(choice["label"]["English(EN)"]),
                        choice["value"],
                    ]
                    rows.append(row)

        return rows

    @classmethod
    def _handle_name_and_label_row(cls, fields: Dict) -> Tuple[List[str], List[str]]:
        names: List[str] = []
        labels: List[str] = []

        for field_name, field_value in fields.items():
            names.append(field_name)
            label = (
                f"{field_value['label']['English(EN)']}"
                f" - {field_value['type']}"
                f"{' - required' if field_value['required'] else ''}"
            )
            labels.append(label)

        return names, labels

    @classmethod
    def _add_template_columns(
        cls, wb: openpyxl.Workbook, business_area_slug: Optional[str] = None, template_for_social_worker: Optional[bool] = None
    ) -> openpyxl.Workbook:
        households_sheet_title = "Households"
        individuals_sheet_title = "Individuals"

        ws_households = wb[households_sheet_title]
        ws_individuals = wb[individuals_sheet_title]
        ws_choices = wb["Choices"]

        flex_fields = serialize_flex_attributes()

        if template_for_social_worker:
            # TODO: ???
            scopes = [Scope.GLOBAL, Scope.XLSX, Scope.XLSX_SOCIAL_WORKER]
        else:
            scopes = [Scope.GLOBAL, Scope.XLSX, Scope.HOUSEHOLD_ID, Scope.COLLECTOR]
        fields = FieldFactory.from_scopes(scopes).apply_business_area(business_area_slug=business_area_slug)

        if not template_for_social_worker:
            households_fields = {
                **fields.associated_with_household().to_dict_by("xlsx_field"),
                **flex_fields[households_sheet_title.lower()],
            }
        else:
            households_fields = {}

        individuals_fields = {
            **fields.associated_with_individual().to_dict_by("xlsx_field"),
            **flex_fields[individuals_sheet_title.lower()],
        }

        households_rows = cls._handle_name_and_label_row(households_fields)
        print("households_rows === ", households_rows)
        individuals_rows = cls._handle_name_and_label_row(individuals_fields)

        for h_row, i_row in zip(households_rows, individuals_rows):
            if not template_for_social_worker:
                ws_households.append(h_row)
            ws_individuals.append(i_row)

        choices = cls._handle_choices({**households_fields, **individuals_fields})
        for row in choices:
            ws_choices.append(row)

        return wb

    @classmethod
    def get_template_file(cls, business_area_slug: Optional[str] = None, template_for_social_worker: Optional[bool] = None) -> openpyxl.Workbook:
        # template_for_social_worker
        # TODO: get columns for social_worker
        return cls._add_template_columns(cls._create_workbook(), business_area_slug, template_for_social_worker)
