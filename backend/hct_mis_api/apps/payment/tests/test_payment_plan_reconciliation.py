import io
import os
import tempfile
from collections import namedtuple
from datetime import timedelta
from decimal import Decimal
from typing import TYPE_CHECKING, Any, Tuple
from unittest.mock import patch
from zipfile import ZipFile

from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone

from openpyxl import load_workbook
from parameterized import parameterized
from pytz import utc

from hct_mis_api.apps.account.fixtures import UserFactory
from hct_mis_api.apps.account.permissions import Permissions
from hct_mis_api.apps.core.base_test_case import APITestCase
from hct_mis_api.apps.core.fixtures import create_afghanistan
from hct_mis_api.apps.core.models import BusinessArea
from hct_mis_api.apps.core.utils import decode_id_string, encode_id_base64
from hct_mis_api.apps.household.fixtures import (
    IndividualRoleInHouseholdFactory,
    create_household_and_individuals,
)
from hct_mis_api.apps.household.models import ROLE_PRIMARY
from hct_mis_api.apps.payment.celery_tasks import (
    create_payment_plan_payment_list_xlsx_per_fsp,
    payment_plan_apply_engine_rule,
)
from hct_mis_api.apps.payment.fixtures import (
    FinancialServiceProviderFactory,
    FspXlsxTemplatePerDeliveryMechanismFactory,
    PaymentFactory,
    PaymentPlanFactory,
    PaymentVerificationFactory,
    PaymentVerificationPlanFactory,
    PaymentVerificationSummaryFactory,
)
from hct_mis_api.apps.payment.models import (
    FinancialServiceProviderXlsxTemplate,
    GenericPayment,
    Payment,
    PaymentPlan,
    PaymentVerification,
    PaymentVerificationPlan,
)
from hct_mis_api.apps.payment.xlsx.xlsx_payment_plan_per_fsp_import_service import (
    XlsxPaymentPlanImportPerFspService,
)
from hct_mis_api.apps.registration_data.fixtures import RegistrationDataImportFactory
from hct_mis_api.apps.steficon.fixtures import RuleCommitFactory, RuleFactory

if TYPE_CHECKING:
    from hct_mis_api.apps.household.models import Household, Individual

CREATE_PROGRAMME_MUTATION = """
mutation CreateProgram($programData: CreateProgramInput!) {
  createProgram(programData: $programData) {
    program {
      id
    }
  }
}
"""

UPDATE_PROGRAM_MUTATION = """
mutation UpdateProgram($programData: UpdateProgramInput!) {
  updateProgram(programData: $programData) {
    program {
      id
    }
  }
}
"""


CREATE_TARGET_POPULATION_MUTATION = """
mutation CreateTP($input: CreateTargetPopulationInput!) {
  createTargetPopulation(input: $input) {
    targetPopulation {
      id
      status
    }
  }
}
"""


CREATE_PAYMENT_PLAN_MUTATION = """
mutation CreatePaymentPlan($input: CreatePaymentPlanInput!) {
    createPaymentPlan(input: $input) {
        paymentPlan {
            id
        }
    }
}
"""

LOCK_TARGET_POPULATION_MUTATION = """
mutation LockTP($id: ID!) {
    lockTargetPopulation(id: $id) {
        targetPopulation {
            id
        }
    }
}
"""

FINALIZE_TARGET_POPULATION_MUTATION = """
mutation FinalizeTP($id: ID!) {
    finalizeTargetPopulation(id: $id) {
        targetPopulation {
            id
            status
        }
    }
}
"""

SET_STEFICON_RULE_MUTATION = """
mutation SetSteficonRuleOnPaymentPlanPaymentList($paymentPlanId: ID!, $steficonRuleId: ID!) {
    setSteficonRuleOnPaymentPlanPaymentList(paymentPlanId: $paymentPlanId, steficonRuleId: $steficonRuleId) {
        paymentPlan {
            id
        }
    }
}
"""


PAYMENT_PLAN_ACTION_MUTATION = """
mutation ActionPaymentPlanMutation($input: ActionPaymentPlanInput!) {
    actionPaymentPlanMutation(input: $input) {
        paymentPlan {
            status
            id
        }
    }
}"""

CHOOSE_DELIVERY_MECHANISMS_MUTATION = """
mutation ChooseDeliveryMechanismsForPaymentPlan($input: ChooseDeliveryMechanismsForPaymentPlanInput!) {
    chooseDeliveryMechanismsForPaymentPlan(input: $input) {
        paymentPlan {
            id
            deliveryMechanisms {
                order
                name
            }
        }
    }
}
"""

AVAILABLE_FSPS_FOR_DELIVERY_MECHANISMS_QUERY = """
query AvailableFspsForDeliveryMechanisms($input: AvailableFspsForDeliveryMechanismsInput!) {
    availableFspsForDeliveryMechanisms(input: $input) {
        deliveryMechanism
        fsps {
            id
            name
        }
    }
}
"""

ASSIGN_FSPS_MUTATION = """
mutation AssignFspToDeliveryMechanism($paymentPlanId: ID!, $mappings: [FSPToDeliveryMechanismMappingInput!]!) {
    assignFspToDeliveryMechanism(input: {
        paymentPlanId: $paymentPlanId,
        mappings: $mappings
    }) {
        paymentPlan {
            id
            deliveryMechanisms {
                name
                order
                fsp {
                    id
                }
            }
        }
    }
}
"""

EXPORT_XLSX_PER_FSP_MUTATION = """
mutation ExportXlsxPaymentPlanPaymentListPerFsp($paymentPlanId: ID!) {
    exportXlsxPaymentPlanPaymentListPerFsp(paymentPlanId: $paymentPlanId) {
        paymentPlan {
            id
        }
    }
}
"""

IMPORT_XLSX_PER_FSP_MUTATION = """
mutation ImportXlsxPaymentPlanPaymentListPerFsp($paymentPlanId: ID!, $file: Upload!) {
    importXlsxPaymentPlanPaymentListPerFsp(paymentPlanId: $paymentPlanId, file: $file) {
        paymentPlan {
            id
        }
        errors {
            sheet
            coordinates
            message
        }
    }
}
"""


class TestPaymentPlanReconciliation(APITestCase):
    @classmethod
    def create_household_and_individual(cls) -> Tuple["Household", "Individual"]:
        household, individuals = create_household_and_individuals(
            household_data={
                "registration_data_import": cls.registration_data_import,
                "business_area": cls.business_area,
            },
            individuals_data=[{}],
        )
        IndividualRoleInHouseholdFactory(household=household, individual=individuals[0], role=ROLE_PRIMARY)
        return household, individuals[0]

    @classmethod
    def setUpTestData(cls) -> None:
        create_afghanistan(
            is_payment_plan_applicable=True,
        )
        cls.business_area = BusinessArea.objects.get(slug="afghanistan")
        cls.user = UserFactory.create()
        cls.all_necessary_permissions = [
            Permissions.PM_CREATE,
            Permissions.PM_VIEW_DETAILS,
            Permissions.PM_VIEW_LIST,
            Permissions.PM_IMPORT_XLSX_WITH_ENTITLEMENTS,
            Permissions.PM_APPLY_RULE_ENGINE_FORMULA_WITH_ENTITLEMENTS,
            Permissions.PROGRAMME_CREATE,
            Permissions.PROGRAMME_ACTIVATE,
            Permissions.TARGETING_CREATE,
            Permissions.TARGETING_LOCK,
            Permissions.TARGETING_SEND,
            Permissions.PM_LOCK_AND_UNLOCK,
            Permissions.PM_LOCK_AND_UNLOCK_FSP,
            Permissions.PM_SEND_FOR_APPROVAL,
            Permissions.PM_ACCEPTANCE_PROCESS_APPROVE,
            Permissions.PM_ACCEPTANCE_PROCESS_AUTHORIZE,
            Permissions.PM_ACCEPTANCE_PROCESS_FINANCIAL_REVIEW,
            Permissions.PM_IMPORT_XLSX_WITH_RECONCILIATION,
        ]
        cls.create_user_role_with_permissions(
            cls.user,
            cls.all_necessary_permissions,
            cls.business_area,
        )

        cls.registration_data_import = RegistrationDataImportFactory(business_area=cls.business_area)

        cls.household_1, cls.individual_1 = cls.create_household_and_individual()
        cls.household_1.refresh_from_db()
        cls.household_2, cls.individual_2 = cls.create_household_and_individual()
        cls.household_3, cls.individual_3 = cls.create_household_and_individual()

    @patch("hct_mis_api.apps.payment.models.PaymentPlan.get_exchange_rate", return_value=2.0)
    def test_receiving_reconciliations_from_fsp(self, mock_get_exchange_rate: Any) -> None:
        create_programme_response = self.graphql_request(
            request_string=CREATE_PROGRAMME_MUTATION,
            context={"user": self.user},
            variables={
                "programData": {
                    "name": "NName",
                    "scope": "UNICEF",
                    "startDate": timezone.datetime(2022, 8, 24, tzinfo=utc),
                    "endDate": timezone.datetime(2022, 8, 31, tzinfo=utc),
                    "description": "desc",
                    "budget": "0.00",
                    "administrativeAreasOfImplementation": "",
                    "populationGoal": 0,
                    "frequencyOfPayments": "REGULAR",
                    "sector": "MULTI_PURPOSE",
                    "cashPlus": True,
                    "individualDataNeeded": False,
                    "businessAreaSlug": self.business_area.slug,
                }
            },
        )
        program_id = create_programme_response["data"]["createProgram"]["program"]["id"]

        self.graphql_request(
            request_string=UPDATE_PROGRAM_MUTATION,
            context={"user": self.user},
            variables={
                "programData": {
                    "id": program_id,
                    "status": "ACTIVE",
                },
            },
        )

        create_target_population_response = self.graphql_request(
            request_string=CREATE_TARGET_POPULATION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "programId": program_id,
                    "name": "TargP",
                    "excludedIds": "",
                    "exclusionReason": "",
                    "businessAreaSlug": self.business_area.slug,
                    "targetingCriteria": {
                        "rules": [
                            {
                                "filters": [
                                    {
                                        "comparisonMethod": "EQUALS",
                                        "arguments": ["True"],
                                        "fieldName": "consent",
                                        "isFlexField": False,
                                    }
                                ],
                                "individualsFiltersBlocks": [],
                            }
                        ]
                    },
                }
            },
        )
        target_population_id = create_target_population_response["data"]["createTargetPopulation"]["targetPopulation"][
            "id"
        ]

        self.graphql_request(
            request_string=LOCK_TARGET_POPULATION_MUTATION,
            context={"user": self.user},
            variables={
                "id": target_population_id,
            },
        )

        finalize_tp_response = self.graphql_request(
            request_string=FINALIZE_TARGET_POPULATION_MUTATION,
            context={"user": self.user},
            variables={
                "id": target_population_id,
            },
        )
        status = finalize_tp_response["data"]["finalizeTargetPopulation"]["targetPopulation"]["status"]
        self.assertEqual(status, "READY_FOR_PAYMENT_MODULE")

        create_payment_plan_response = self.graphql_request(
            request_string=CREATE_PAYMENT_PLAN_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "businessAreaSlug": self.business_area.slug,
                    "targetingId": target_population_id,
                    "startDate": timezone.datetime(2022, 8, 24, tzinfo=utc),
                    "endDate": timezone.datetime(2022, 8, 31, tzinfo=utc),
                    "dispersionStartDate": (timezone.now() - timedelta(days=1)).strftime("%Y-%m-%d"),
                    "dispersionEndDate": (timezone.now() + timedelta(days=1)).strftime("%Y-%m-%d"),
                    "currency": "USD",
                }
            },
        )
        assert "errors" not in create_payment_plan_response, create_payment_plan_response
        encoded_payment_plan_id = create_payment_plan_response["data"]["createPaymentPlan"]["paymentPlan"]["id"]
        payment_plan_id = decode_id_string(encoded_payment_plan_id)

        santander_fsp = FinancialServiceProviderFactory(
            name="Santander",
            delivery_mechanisms=[GenericPayment.DELIVERY_TYPE_CASH, GenericPayment.DELIVERY_TYPE_TRANSFER],
            distribution_limit=None,
        )
        FspXlsxTemplatePerDeliveryMechanismFactory(
            financial_service_provider=santander_fsp, delivery_mechanism=GenericPayment.DELIVERY_TYPE_CASH
        )
        FspXlsxTemplatePerDeliveryMechanismFactory(
            financial_service_provider=santander_fsp, delivery_mechanism=GenericPayment.DELIVERY_TYPE_TRANSFER
        )
        encoded_santander_fsp_id = encode_id_base64(santander_fsp.id, "FinancialServiceProvider")

        payment = PaymentFactory(
            parent=PaymentPlan.objects.get(id=payment_plan_id),
            business_area=self.business_area,
            household=self.household_1,
            collector=self.individual_1,
            delivery_type=None,
            entitlement_quantity=1000,
            entitlement_quantity_usd=100,
            delivered_quantity=None,
            delivered_quantity_usd=None,
            financial_service_provider=None,
            excluded=False,
        )
        self.assertEqual(payment.entitlement_quantity, 1000)

        lock_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "LOCK",
                }
            },
        )
        assert "errors" not in lock_payment_plan_response, lock_payment_plan_response
        assert lock_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"] == "LOCKED"

        rule = RuleFactory(name="Rule")
        RuleCommitFactory(definition="result.value=Decimal('500')", rule=rule)

        payment_plan = PaymentPlan.objects.get(id=payment_plan_id)
        self.assertEqual(payment_plan.background_action_status, None)

        with patch("hct_mis_api.apps.payment.mutations.payment_plan_apply_engine_rule") as mock:
            set_steficon_response = self.graphql_request(
                request_string=SET_STEFICON_RULE_MUTATION,
                context={"user": self.user},
                variables={
                    "paymentPlanId": encoded_payment_plan_id,
                    "steficonRuleId": encode_id_base64(rule.id, "Rule"),
                },
            )
            assert "errors" not in set_steficon_response, set_steficon_response
            assert mock.delay.call_count == 1
            call_args = mock.delay.call_args[0]
            payment_plan_apply_engine_rule(*call_args)

        payment_plan.refresh_from_db()
        self.assertEqual(payment_plan.background_action_status, None)

        payment.refresh_from_db()
        self.assertEqual(payment.entitlement_quantity, 500)

        choose_dms_response = self.graphql_request(
            request_string=CHOOSE_DELIVERY_MECHANISMS_MUTATION,
            context={"user": self.user},
            variables=dict(
                input=dict(
                    paymentPlanId=encoded_payment_plan_id,
                    deliveryMechanisms=[
                        GenericPayment.DELIVERY_TYPE_CASH,
                    ],
                )
            ),
        )
        assert "errors" not in choose_dms_response, choose_dms_response

        available_fsps_query_response = self.graphql_request(
            request_string=AVAILABLE_FSPS_FOR_DELIVERY_MECHANISMS_QUERY,
            context={"user": self.user},
            variables=dict(
                input={
                    "paymentPlanId": encoded_payment_plan_id,
                }
            ),
        )
        assert "errors" not in available_fsps_query_response, available_fsps_query_response
        available_fsps_data = available_fsps_query_response["data"]["availableFspsForDeliveryMechanisms"]
        assert len(available_fsps_data) == 1
        fsps = available_fsps_data[0]["fsps"]
        assert len(fsps) > 0
        assert fsps[0]["name"] == santander_fsp.name

        assign_fsp_mutation_response = self.graphql_request(
            request_string=ASSIGN_FSPS_MUTATION,
            context={"user": self.user},
            variables={
                "paymentPlanId": encoded_payment_plan_id,
                "mappings": [
                    {
                        "deliveryMechanism": GenericPayment.DELIVERY_TYPE_CASH,
                        "fspId": encoded_santander_fsp_id,
                        "order": 1,
                    }
                ],
            },
        )
        assert "errors" not in assign_fsp_mutation_response, assign_fsp_mutation_response

        lock_fsp_in_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "LOCK_FSP",
                }
            },
        )
        assert "errors" not in lock_fsp_in_payment_plan_response, lock_fsp_in_payment_plan_response
        self.assertEqual(
            lock_fsp_in_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "LOCKED_FSP",
        )

        payment_plan.refresh_from_db()
        assert (
            payment_plan.delivery_mechanisms.filter(
                financial_service_provider=santander_fsp, delivery_mechanism=GenericPayment.DELIVERY_TYPE_CASH
            ).count()
            == 1
        )
        assert (
            payment_plan.not_excluded_payments.filter(
                financial_service_provider__isnull=False,
                delivery_type__isnull=False,
            ).count()
            == 1
        )

        send_for_approval_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "SEND_FOR_APPROVAL",
                }
            },
        )
        assert "errors" not in send_for_approval_payment_plan_response, send_for_approval_payment_plan_response
        self.assertEqual(
            send_for_approval_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "IN_APPROVAL",
        )

        approve_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "APPROVE",
                }
            },
        )
        assert "errors" not in approve_payment_plan_response, approve_payment_plan_response
        self.assertEqual(
            approve_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "IN_AUTHORIZATION",
        )

        authorize_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "AUTHORIZE",
                }
            },
        )
        assert "errors" not in authorize_payment_plan_response, authorize_payment_plan_response
        self.assertEqual(
            authorize_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "IN_REVIEW",
        )

        review_payment_plan_response = self.graphql_request(
            request_string=PAYMENT_PLAN_ACTION_MUTATION,
            context={"user": self.user},
            variables={
                "input": {
                    "paymentPlanId": encoded_payment_plan_id,
                    "action": "REVIEW",
                }
            },
        )
        assert "errors" not in review_payment_plan_response, review_payment_plan_response
        self.assertEqual(
            review_payment_plan_response["data"]["actionPaymentPlanMutation"]["paymentPlan"]["status"],
            "ACCEPTED",
        )

        payment_plan.refresh_from_db()
        self.assertEqual(payment_plan.background_action_status, None)

        with patch(
            "hct_mis_api.apps.payment.services.payment_plan_services.create_payment_plan_payment_list_xlsx_per_fsp"
        ) as mock_export:
            export_file_mutation = self.graphql_request(
                request_string=EXPORT_XLSX_PER_FSP_MUTATION,
                context={"user": self.user},
                variables={
                    "paymentPlanId": encoded_payment_plan_id,
                },
            )
            assert "errors" not in export_file_mutation, export_file_mutation
            assert mock_export.delay.call_count == 1
            call_args = mock_export.delay.call_args[0]
            create_payment_plan_payment_list_xlsx_per_fsp(*call_args)

        payment_plan.refresh_from_db()
        zip_file = payment_plan.export_file
        assert zip_file is not None

        with tempfile.TemporaryDirectory() as temp_dir:
            self.assertEqual(len(os.listdir(temp_dir)), 0)

            assert zip_file.file is not None
            with ZipFile(zip_file.file, "r") as zip_ref:
                self.assertEqual(len(zip_ref.namelist()), 1)
                zip_ref.extractall(temp_dir)

            self.assertEqual(len(os.listdir(temp_dir)), 1)

            file_name = os.listdir(temp_dir)[0]
            assert file_name.endswith(".xlsx")
            file_path = os.path.join(temp_dir, file_name)

            workbook = load_workbook(file_path)
            assert workbook.sheetnames == ["Santander"], workbook.sheetnames

            sheet = workbook["Santander"]
            assert sheet.max_row == 2, sheet.max_row

            self.assertEqual(sheet.cell(row=1, column=1).value, "payment_id")
            assert payment_plan.payment_items.count() == 1
            payment = payment_plan.payment_items.first()
            self.assertEqual(sheet.cell(row=2, column=1).value, payment.unicef_id)  # unintuitive

            self.assertEqual(sheet.cell(row=1, column=2).value, "household_id")
            self.assertEqual(sheet.cell(row=2, column=2).value, self.household_1.unicef_id)
            self.assertEqual(sheet.cell(row=1, column=3).value, "household_size")
            self.assertEqual(sheet.cell(row=2, column=3).value, self.household_1.size)
            self.assertEqual(sheet.cell(row=1, column=4).value, "collector_name")
            self.assertEqual(sheet.cell(row=2, column=4).value, payment.collector.full_name)
            self.assertEqual(sheet.cell(row=1, column=5).value, "payment_channel")
            self.assertEqual(sheet.cell(row=2, column=5).value, "Cash")
            self.assertEqual(sheet.cell(row=1, column=6).value, "fsp_name")
            self.assertEqual(sheet.cell(row=2, column=6).value, payment.financial_service_provider.name)
            self.assertEqual(sheet.cell(row=1, column=7).value, "currency")
            self.assertEqual(sheet.cell(row=2, column=7).value, payment.currency)
            self.assertEqual(sheet.cell(row=1, column=8).value, "entitlement_quantity")
            self.assertEqual(sheet.cell(row=2, column=8).value, payment.entitlement_quantity)
            self.assertEqual(sheet.cell(row=1, column=9).value, "entitlement_quantity_usd")
            self.assertEqual(sheet.cell(row=2, column=9).value, payment.entitlement_quantity_usd)
            self.assertEqual(sheet.cell(row=1, column=10).value, "delivered_quantity")
            self.assertEqual(sheet.cell(row=2, column=10).value, None)

            payment.refresh_from_db()
            self.assertEqual(payment.entitlement_quantity, 500)
            self.assertEqual(payment.delivered_quantity, None)
            self.assertEqual(payment.status, Payment.STATUS_PENDING)
            self.assertEqual(payment_plan.is_reconciled, False)

            filled_file_name = "filled.xlsx"
            filled_file_path = os.path.join(temp_dir, filled_file_name)

            # update xls, delivered_quantity != entitlement_quantity
            sheet.cell(
                row=2, column=FinancialServiceProviderXlsxTemplate.DEFAULT_COLUMNS.index("delivered_quantity") + 1
            ).value = 666
            workbook.save(filled_file_path)

            with open(filled_file_path, "rb") as file:
                uploaded_file = SimpleUploadedFile(filled_file_name, file.read())
                with patch("hct_mis_api.apps.payment.services.payment_plan_services.transaction") as mock_import:
                    import_mutation_response = self.graphql_request(
                        request_string=IMPORT_XLSX_PER_FSP_MUTATION,
                        context={"user": self.user},
                        variables={
                            "paymentPlanId": encoded_payment_plan_id,
                            "file": uploaded_file,
                        },
                    )
                    assert (
                        "errors" in import_mutation_response["data"]["importXlsxPaymentPlanPaymentListPerFsp"]
                    ), import_mutation_response
                    assert (
                        import_mutation_response["data"]["importXlsxPaymentPlanPaymentListPerFsp"]["errors"][0][
                            "message"
                        ]
                        == f"Payment {payment.unicef_id}: Delivered quantity 666.00 is bigger than Entitlement quantity 500.00"
                    ), import_mutation_response

            # update xls, delivered_quantity == entitlement_quantity
            sheet.cell(
                row=2, column=FinancialServiceProviderXlsxTemplate.DEFAULT_COLUMNS.index("delivered_quantity") + 1
            ).value = payment.entitlement_quantity
            workbook.save(filled_file_path)

            with open(filled_file_path, "rb") as file:
                uploaded_file = SimpleUploadedFile(filled_file_name, file.read())
                with patch("hct_mis_api.apps.payment.services.payment_plan_services.transaction") as mock_import:
                    import_mutation_response = self.graphql_request(
                        request_string=IMPORT_XLSX_PER_FSP_MUTATION,
                        context={"user": self.user},
                        variables={
                            "paymentPlanId": encoded_payment_plan_id,
                            "file": uploaded_file,
                        },
                    )
                    assert (
                        import_mutation_response["data"]["importXlsxPaymentPlanPaymentListPerFsp"]["errors"] is None
                    ), import_mutation_response
                    assert mock_import.on_commit.call_count == 1
                    mock_import.on_commit.call_args[0][0]()  # call real func

            payment.refresh_from_db()
            payment.household.refresh_from_db()
            self.assertEqual(payment.entitlement_quantity, 500)
            self.assertEqual(payment.delivered_quantity, 500)
            self.assertEqual(payment.status, Payment.STATUS_DISTRIBUTION_SUCCESS)
            self.assertEqual(payment.household.total_cash_received, 500)
            self.assertEqual(payment.household.total_cash_received_usd, 250)
            self.assertEqual(payment_plan.is_reconciled, True)

    @parameterized.expand(
        [
            (-1, None, Payment.STATUS_ERROR),
            (0, Decimal(0), Payment.STATUS_NOT_DISTRIBUTED),
            (400, Decimal(400), Payment.STATUS_DISTRIBUTION_PARTIAL),
            (500, Decimal(500), Payment.STATUS_DISTRIBUTION_SUCCESS),
            (600, None, None),
        ]
    )
    def test_receiving_payment_reconciliations_status(
        self, delivered_quantity: float, expected_delivered_quantity: Decimal, expected_status: str
    ) -> None:
        service = XlsxPaymentPlanImportPerFspService(PaymentPlanFactory(), None)  # type: ignore

        if not expected_status:
            with self.assertRaisesMessage(
                service.XlsxPaymentPlanImportPerFspServiceException,
                "Invalid delivered_quantity 600 provided for payment_id xx",
            ):
                service._get_delivered_quantity_status_and_value(delivered_quantity, Decimal(500), "xx")

        else:
            status, value = service._get_delivered_quantity_status_and_value(delivered_quantity, Decimal(500), "xx")
            self.assertEqual(status, expected_status)
            self.assertEqual(value, expected_delivered_quantity)

    def test_xlsx_payment_plan_import_per_fsp_service_import_row(self) -> None:
        pp = PaymentPlanFactory(status=PaymentPlan.Status.FINISHED)
        pp.refresh_from_db()
        PaymentVerificationSummaryFactory(generic_fk_obj=pp)
        pvp = PaymentVerificationPlanFactory(
            generic_fk_obj=pp,
            verification_channel=PaymentVerificationPlan.VERIFICATION_CHANNEL_MANUAL,
            status=PaymentVerificationPlan.STATUS_ACTIVE,
        )

        payment_1 = PaymentFactory(
            parent=PaymentPlan.objects.get(id=pp.id),
            business_area=self.business_area,
            household=self.household_1,
            collector=self.individual_1,
            delivery_type=None,
            entitlement_quantity=1111,
            entitlement_quantity_usd=100,
            delivered_quantity=1000,
            delivered_quantity_usd=99,
            financial_service_provider=None,
            excluded=False,
        )
        payment_2 = PaymentFactory(
            parent=PaymentPlan.objects.get(id=pp.id),
            business_area=self.business_area,
            household=self.household_2,
            collector=self.individual_2,
            delivery_type=None,
            entitlement_quantity=2222,
            entitlement_quantity_usd=100,
            delivered_quantity=2000,
            delivered_quantity_usd=500,
            financial_service_provider=None,
            excluded=False,
        )
        verification_1 = PaymentVerificationFactory(
            payment_verification_plan=pvp,
            generic_fk_obj=payment_1,
            status=PaymentVerification.STATUS_RECEIVED_WITH_ISSUES,
            received_amount=999,
        )
        verification_2 = PaymentVerificationFactory(
            payment_verification_plan=pvp,
            generic_fk_obj=payment_2,
            status=PaymentVerification.STATUS_RECEIVED,
            received_amount=500,
        )
        import_xlsx_service = XlsxPaymentPlanImportPerFspService(pp, io.BytesIO())
        import_xlsx_service.xlsx_headers = ["payment_id", "delivered_quantity"]
        import_xlsx_service.payments_dict[str(payment_1.pk)] = payment_1
        import_xlsx_service.payments_dict[str(payment_2.pk)] = payment_2

        row = namedtuple(
            "row",
            [
                "value",
            ],
        )

        import_xlsx_service._import_row([row(str(payment_1.id)), row(999)], 1)
        import_xlsx_service._import_row([row(str(payment_2.id)), row(100)], 1)
        payment_1.save()
        payment_2.save()
        # Update payment Verification
        PaymentVerification.objects.bulk_update(
            import_xlsx_service.payment_verifications_to_save, ("status", "status_date")
        )
        payment_1.refresh_from_db()
        payment_2.refresh_from_db()
        verification_1.refresh_from_db()
        verification_2.refresh_from_db()

        self.assertEqual(payment_1.delivered_quantity, 999)
        self.assertEqual(verification_1.received_amount, 999)
        self.assertEqual(verification_1.status, PaymentVerification.STATUS_RECEIVED)

        self.assertEqual(payment_2.delivered_quantity, 100)
        self.assertEqual(verification_2.received_amount, 500)
        self.assertEqual(verification_2.status, PaymentVerification.STATUS_RECEIVED_WITH_ISSUES)
