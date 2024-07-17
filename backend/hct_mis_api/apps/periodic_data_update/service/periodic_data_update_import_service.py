from typing import Any, Optional, Union

from django import forms
from django.core.exceptions import ValidationError
from django.core.files import File
from django.db import transaction

import openpyxl

from hct_mis_api.apps.core.models import FlexibleAttribute, PeriodicFieldData
from hct_mis_api.apps.household.models import Individual
from hct_mis_api.apps.periodic_data_update.models import (
    PeriodicDataUpdateTemplate,
    PeriodicDataUpdateUpload,
)
from hct_mis_api.apps.periodic_data_update.service.periodic_data_update_export_template_service import (
    PeriodicDataUpdateExportTemplateService,
)


class PeriodicDataUpdateBaseForm(forms.Form):
    individual__uuid = forms.UUIDField()
    individual_unicef_id = forms.CharField()
    first_name = forms.CharField()
    last_name = forms.CharField()


class PeriodicDataUpdateImportService:
    def __init__(self, periodic_data_update_upload: PeriodicDataUpdateUpload) -> None:
        self.periodic_data_update_upload = periodic_data_update_upload
        self.periodic_data_update_template = self.periodic_data_update_upload.template
        self.file = self.periodic_data_update_upload.file

    def import_data(self) -> None:
        try:
            with transaction.atomic():
                self._open_workbook()
                self._read_flexible_attributes()
                cleaned_data_list = self._read_rows()
                self._update_individuals(cleaned_data_list)
                self.file.close()
                self.periodic_data_update_upload.status = PeriodicDataUpdateUpload.Status.SUCCESSFUL
                self.periodic_data_update_upload.save()
        except ValidationError as e:
            raise e
            self.periodic_data_update_upload.status = PeriodicDataUpdateUpload.Status.FAILED
            self.periodic_data_update_upload.error_message = str(e)
            self.periodic_data_update_upload.save()
            return

    def _open_workbook(self) -> None:
        self.wb = openpyxl.load_workbook(self.file)
        self.ws_pdu = self.wb[PeriodicDataUpdateExportTemplateService.PDU_SHEET]
        self.ws_meta = self.wb[PeriodicDataUpdateExportTemplateService.META_SHEET]
        self.periodic_data_update_template_id: Optional[PeriodicDataUpdateTemplate] = None
        self.flexible_attributes_dict: Optional[dict[str, FlexibleAttribute]] = None

    @classmethod
    def read_periodic_data_update_template_object(cls, file: File) -> PeriodicDataUpdateTemplate:
        wb = openpyxl.load_workbook(file)  # type: ignore
        ws_meta = wb[PeriodicDataUpdateExportTemplateService.META_SHEET]
        try:
            periodic_data_update_template_id = wb.custom_doc_props[
                PeriodicDataUpdateExportTemplateService.PROPERTY_ID_NAME
            ]
        except KeyError:
            periodic_data_update_template_id = None
        if periodic_data_update_template_id:
            periodic_data_update_template_id = periodic_data_update_template_id.value
        if not periodic_data_update_template_id:
            periodic_data_update_template_id = ws_meta[PeriodicDataUpdateExportTemplateService.META_ID_ADDRESS].value
        if not periodic_data_update_template_id:
            raise ValidationError("Periodic Data Update Template ID is missing in the file")
        try:
            if isinstance(periodic_data_update_template_id, str):
                periodic_data_update_template_id = periodic_data_update_template_id.strip()
                periodic_data_update_template_id = int(periodic_data_update_template_id)
        except ValueError:
            raise ValidationError("Periodic Data Update Template ID must be a number")
        if type(periodic_data_update_template_id) is not int:
            raise ValidationError("Periodic Data Update Template ID must be an integer")

        periodic_data_update_template = PeriodicDataUpdateTemplate.objects.filter(
            id=periodic_data_update_template_id
        ).first()
        if not periodic_data_update_template:
            raise ValidationError(f"Periodic Data Update Template with ID {periodic_data_update_template_id} not found")
        return periodic_data_update_template

    def _read_flexible_attributes(self) -> None:
        rounds_data = self.periodic_data_update_template.rounds_data
        fields_name_list = [field["field"] for field in rounds_data]
        fields_name = set(fields_name_list)
        flexible_attributes = FlexibleAttribute.objects.filter(name__in=fields_name, type=FlexibleAttribute.PDU)
        if len(flexible_attributes) != len(fields_name):
            raise ValidationError("Some fields are missing in the flexible attributes")
        self.flexible_attributes_dict = {field.name: field for field in flexible_attributes}

    def _read_header(self) -> list[str]:
        header = [cell.value for cell in self.ws_pdu[1]]
        return header

    def _read_rows(self) -> list[dict]:
        header = self._read_header()
        errors = []
        cleaned_data_list = []
        for row in self.ws_pdu.iter_rows(min_row=2, values_only=True):
            cleaned_data = self._read_row(errors, header, row)
            if not cleaned_data:
                continue
            cleaned_data_list.append(cleaned_data)
        if errors:
            raise ValidationError(errors)
        return cleaned_data_list

    def _read_row(self, errors: list, header: list, row: list) -> Optional[dict]:
        row_empty_values = []
        for value in row:
            if value == "-":
                row_empty_values.append(None)
            else:
                row_empty_values.append(value)
        data = dict(zip(header, row_empty_values))
        form = self._build_form()(data=data)
        if not form.is_valid():
            errors.append(form.errors)
            return None
        cleaned_data = form.cleaned_data
        return cleaned_data

    def _update_individuals(self, cleaned_data_list: list[dict]) -> None:
        individuals = []
        for cleaned_data in cleaned_data_list:
            individual = self._import_cleaned_data(cleaned_data)
            individuals.append(individual)
        Individual.objects.bulk_update(individuals, ["flex_fields"])

    def _import_cleaned_data(self, cleaned_data: dict) -> Individual:
        individual_uuid = cleaned_data["individual__uuid"]
        individual_unicef_id = cleaned_data["individual_unicef_id"]
        individual = Individual.objects.filter(id=individual_uuid).first()
        for round in self.periodic_data_update_template.rounds_data:
            field_name = round["field"]
            round_number = round["round"]
            round_number_from_xlsx = cleaned_data[f"{field_name}__round_number"]
            value_from_xlsx = cleaned_data[f"{field_name}__round_value"]
            collection_date_from_xlsx = cleaned_data[f"{field_name}__collection_date"]
            if value_from_xlsx is None:
                continue
            if round_number_from_xlsx != round_number:
                raise ValidationError(
                    f"Round number mismatch for field {field_name} and individual {individual_uuid} / {individual_unicef_id}"
                )
            if not individual:
                raise ValidationError(f"Individual with UUID {individual_uuid} / {individual_unicef_id} not found")
            current_value = self._get_round_value(individual, field_name, round_number)
            if current_value and value_from_xlsx:
                raise ValidationError(
                    f"Value already exists for field {field_name} for round {round_number} and individual {individual_uuid} / {individual_unicef_id}"
                )
            self._set_round_value(individual, field_name, round_number, value_from_xlsx, collection_date_from_xlsx)
        return individual

    def _get_round_value(
        self, individual: Individual, pdu_field_name: str, round_number: int
    ) -> Optional[Union[str, int, float, bool]]:
        flex_fields_data = individual.flex_fields
        field_data = flex_fields_data.get(pdu_field_name)
        if field_data:
            round_data = field_data.get(str(round_number))
            if round_data:
                return round_data.get("value")
        return None

    def _set_round_value(
        self, individual: Individual, pdu_field_name: str, round_number: int, value: Any, collection_date: Any
    ) -> None:
        flex_fields_data = individual.flex_fields
        if pdu_field_name not in flex_fields_data:
            flex_fields_data[pdu_field_name] = {}
        field_data = flex_fields_data[pdu_field_name]
        if str(round_number) not in field_data:
            field_data[str(round_number)] = {}
        round_data = field_data.get(str(round_number))
        round_data["value"] = value
        round_data["collection_date"] = collection_date

    def _build_form(self) -> type[forms.Form]:
        form_fields_dict: dict[str, forms.Field] = {}
        for round in self.periodic_data_update_template.rounds_data:
            flexible_attribute = self.flexible_attributes_dict.get(round["field"])
            if not flexible_attribute:
                raise ValidationError(f"Flexible Attribute for field {round['field']} not found")
            form_fields_dict[f"{round['field']}__round_number"] = forms.IntegerField()
            form_fields_dict[f"{round['field']}__round_name"] = forms.CharField(required=False)
            form_fields_dict[f"{round['field']}__round_value"] = self._get_form_field_for_value(flexible_attribute)
            form_fields_dict[f"{round['field']}__collection_date"] = forms.DateField(required=False)

        return type("PeriodicDataUpdateForm", (PeriodicDataUpdateBaseForm,), form_fields_dict)

    def _get_form_field_for_value(self, flexible_attribute: FlexibleAttribute) -> forms.Field:
        if flexible_attribute.pdu_data.subtype == PeriodicFieldData.STRING:
            return forms.CharField(required=False)
        elif flexible_attribute.pdu_data.subtype == PeriodicFieldData.DECIMAL:
            return forms.DecimalField(required=False)
        elif flexible_attribute.pdu_data.subtype == PeriodicFieldData.BOOLEAN:
            return forms.BooleanField(required=False)
        elif flexible_attribute.pdu_data.subtype == PeriodicFieldData.DATE:
            return forms.DateField(required=False)
        raise ValidationError(f"Invalid subtype for field {flexible_attribute.name}")
